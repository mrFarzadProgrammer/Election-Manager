from __future__ import annotations

import io
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

import auth
import database
import models
import schemas


router = APIRouter(tags=["monitoring"])


def _parse_iso_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        s = str(value).strip()
        if not s:
            return None
        if "T" not in s:
            return datetime.fromisoformat(s + "T00:00:00")
        return datetime.fromisoformat(s.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def _log_export_action(db: Session, *, admin_id: int, export_type: str, filters: dict | None = None) -> None:
    try:
        db.add(models.AdminExportLog(admin_id=int(admin_id), export_type=str(export_type), filters=filters or None))
        db.commit()
    except Exception:
        db.rollback()


@router.get("/api/admin/monitoring/errors", response_model=list[schemas.TechnicalErrorItem])
def admin_monitoring_errors(
    representative_id: int | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    limit: int = 500,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    limit = max(1, min(int(limit), 5000))
    q = db.query(models.TechnicalErrorLog)
    if representative_id is not None:
        q = q.filter(models.TechnicalErrorLog.candidate_id == int(representative_id))
    sd = _parse_iso_dt(start_date)
    ed = _parse_iso_dt(end_date)
    if sd is not None:
        q = q.filter(models.TechnicalErrorLog.created_at >= sd)
    if ed is not None:
        q = q.filter(models.TechnicalErrorLog.created_at <= ed)
    return q.order_by(models.TechnicalErrorLog.id.desc()).limit(limit).all()


@router.get("/api/admin/monitoring/errors/export.xlsx")
def admin_monitoring_errors_export_xlsx(
    representative_id: int | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    rows = admin_monitoring_errors(
        representative_id=representative_id,
        start_date=start_date,
        end_date=end_date,
        limit=5000,
        db=db,
        current_user=current_user,
    )
    _log_export_action(
        db,
        admin_id=int(current_user.id),
        export_type="technical_errors",
        filters={"representative_id": representative_id, "start_date": start_date, "end_date": end_date},
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "errors"
    ws.append(["error_id", "timestamp", "service_name", "error_type", "error_message", "user_id", "representative_id", "state"])
    for r in rows:
        ws.append(
            [
                int(getattr(r, "id", 0) or 0),
                getattr(r, "created_at", None),
                str(getattr(r, "service_name", "") or ""),
                str(getattr(r, "error_type", "") or ""),
                str(getattr(r, "error_message", "") or ""),
                str(getattr(r, "telegram_user_id", "") or ""),
                getattr(r, "candidate_id", None),
                str(getattr(r, "state", "") or ""),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"monitoring_errors_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/api/admin/monitoring/ux-logs", response_model=list[schemas.MonitoringUxLogItem])
def admin_monitoring_ux_logs(
    representative_id: int | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    action: str | None = None,
    limit: int = 1000,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    limit = max(1, min(int(limit), 5000))
    q = db.query(models.BotUxLog)
    if representative_id is not None:
        q = q.filter(models.BotUxLog.candidate_id == int(representative_id))
    sd = _parse_iso_dt(start_date)
    ed = _parse_iso_dt(end_date)
    if sd is not None:
        q = q.filter(models.BotUxLog.created_at >= sd)
    if ed is not None:
        q = q.filter(models.BotUxLog.created_at <= ed)
    if action:
        q = q.filter(models.BotUxLog.action == str(action).strip())
    return q.order_by(models.BotUxLog.id.desc()).limit(limit).all()


@router.get("/api/admin/monitoring/ux-logs/export.xlsx")
def admin_monitoring_ux_logs_export_xlsx(
    representative_id: int | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    action: str | None = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    rows = admin_monitoring_ux_logs(
        representative_id=representative_id,
        start_date=start_date,
        end_date=end_date,
        action=action,
        limit=5000,
        db=db,
        current_user=current_user,
    )
    _log_export_action(
        db,
        admin_id=int(current_user.id),
        export_type="ux_logs",
        filters={"representative_id": representative_id, "start_date": start_date, "end_date": end_date, "action": action},
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "ux_logs"
    ws.append(["log_id", "timestamp", "user_id", "representative_id", "current_state", "action", "expected_action"])
    for r in rows:
        ws.append(
            [
                int(getattr(r, "id", 0) or 0),
                getattr(r, "created_at", None),
                str(getattr(r, "telegram_user_id", "") or ""),
                int(getattr(r, "candidate_id", 0) or 0),
                str(getattr(r, "state", "") or ""),
                str(getattr(r, "action", "") or ""),
                str(getattr(r, "expected_action", "") or ""),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"monitoring_ux_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/api/admin/monitoring/flow-drops", response_model=list[schemas.FlowDropItem])
def admin_monitoring_flow_drops(
    representative_id: int | None = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    q = db.query(models.BotFlowDropCounter)
    if representative_id is not None:
        q = q.filter(models.BotFlowDropCounter.candidate_id == int(representative_id))
    return q.order_by(models.BotFlowDropCounter.updated_at.desc()).all()


@router.get("/api/admin/monitoring/flow-drops/export.xlsx")
def admin_monitoring_flow_drops_export_xlsx(
    representative_id: int | None = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    rows = admin_monitoring_flow_drops(representative_id=representative_id, db=db, current_user=current_user)
    _log_export_action(
        db,
        admin_id=int(current_user.id),
        export_type="flow_drop_stats",
        filters={"representative_id": representative_id},
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "flow_drops"
    ws.append(["id", "representative_id", "flow_type", "started_count", "completed_count", "abandoned_count", "updated_at"])
    for r in rows:
        ws.append(
            [
                int(getattr(r, "id", 0) or 0),
                getattr(r, "candidate_id", None),
                str(getattr(r, "flow_type", "") or ""),
                int(getattr(r, "started_count", 0) or 0),
                int(getattr(r, "completed_count", 0) or 0),
                int(getattr(r, "abandoned_count", 0) or 0),
                getattr(r, "updated_at", None),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"monitoring_flow_drops_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/api/admin/monitoring/health-checks", response_model=list[schemas.HealthCheckItem])
def admin_monitoring_health_checks(
    representative_id: int | None = None,
    check_type: str | None = None,
    limit: int = 500,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    limit = max(1, min(int(limit), 5000))
    q = db.query(models.BotHealthCheck)
    if representative_id is not None:
        q = q.filter(models.BotHealthCheck.candidate_id == int(representative_id))
    if check_type:
        q = q.filter(models.BotHealthCheck.check_type == str(check_type).strip())
    return q.order_by(models.BotHealthCheck.id.desc()).limit(limit).all()
