# main.py
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from sqlalchemy import func, case
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel
from typing import List, Optional
from jose import jwt
from contextlib import asynccontextmanager
import models, database, auth, schemas
import os
import shutil
import re
import logging
import mimetypes
import httpx
import io
import json
from dotenv import load_dotenv
import jdatetime
from datetime import datetime, timedelta, timezone
import uuid

from openpyxl import Workbook

load_dotenv()

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Runtime config (dev vs prod)
# ---------------------------------------------------------------------------

APP_ENV = (os.getenv("APP_ENV") or os.getenv("ENV") or "development").strip().lower()


def _env_truthy(name: str) -> bool:
    return (os.getenv(name) or "").strip().lower() in {"1", "true", "yes", "on"}


def _client_ip(request: Request) -> str:
    # Behind reverse proxy, enable TRUST_PROXY=1 and ensure X-Forwarded-For is set.
    if _env_truthy("TRUST_PROXY"):
        xff = (request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
        if xff:
            return xff
    try:
        return request.client.host if request.client else "unknown"
    except Exception:
        return "unknown"


# Very small in-memory rate limiter (per-process). Use Redis-based limiter for multi-worker production.
_RATE_LIMIT: dict[str, list[float]] = {}


def _rate_limit(request: Request, *, key: str, limit: int, window_seconds: int) -> None:
    now = datetime.now(timezone.utc).timestamp()
    bucket_key = f"{key}:{_client_ip(request)}"
    items = _RATE_LIMIT.get(bucket_key, [])
    cutoff = now - window_seconds
    items = [t for t in items if t >= cutoff]
    if len(items) >= limit:
        raise HTTPException(status_code=429, detail="تعداد درخواست‌ها زیاد است. لطفاً کمی بعد دوباره تلاش کنید.")
    items.append(now)
    _RATE_LIMIT[bucket_key] = items


# ============================================================================
# Commitments (Public Digital Contracts) - STRICT RULES
# ============================================================================

COMMITMENT_TERMS_VERSION = "v1"

ALLOWED_COMMITMENT_STATUSES = {"draft", "active", "in_progress", "completed", "failed"}
ALLOWED_COMMITMENT_STATUS_AFTER_PUBLISH = {"active", "in_progress", "completed", "failed"}
ALLOWED_COMMITMENT_CATEGORIES = {"economy", "housing", "transparency", "employment", "other"}


def _log_commitment_security_event(
    *,
    db: Session,
    representative_id: int,
    error_type: str,
    message: str,
) -> None:
    try:
        db.add(
            models.TechnicalErrorLog(
                service_name="api",
                error_type=str(error_type),
                error_message=str(message)[:8000],
                telegram_user_id=None,
                candidate_id=int(representative_id),
                state="commitments",
            )
        )
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass


def _require_candidate(current_user: models.User) -> None:
    if current_user.role != "CANDIDATE":
        raise HTTPException(status_code=403, detail="Access denied")


def _normalize_commitment_status(value: str | None) -> str:
    return (value or "").strip().lower().replace(" ", "_")


def _validate_commitment_category(value: str | None) -> str:
    v = (value or "").strip().lower()
    if v not in ALLOWED_COMMITMENT_CATEGORIES:
        raise HTTPException(status_code=422, detail={"message": "Invalid category"})
    return v


def _validate_commitment_status_after_publish(value: str | None) -> str:
    v = _normalize_commitment_status(value)
    if v not in ALLOWED_COMMITMENT_STATUS_AFTER_PUBLISH:
        raise HTTPException(status_code=422, detail={"message": "Invalid status"})
    return v


def _extract_telegram_chat_target(value: str | None) -> str | int | None:
    """Best-effort normalization for Telegram chat targets.

    Supports:
    - numeric chat ids (e.g. -100123...)
    - @username
    - https://t.me/username

    Returns None for invite links (t.me/+...) and unknown formats.
    """
    raw = (value or "").strip()
    if not raw:
        return None

    # Numeric chat id
    if re.fullmatch(r"-?\d{5,}", raw):
        try:
            return int(raw)
        except Exception:
            return raw

    v = raw
    if v.startswith("t.me/"):
        v = "https://" + v

    if v.startswith("http://") or v.startswith("https://"):
        m = re.search(r"t\.me/([^/?#]+)", v)
        if not m:
            return None
        v = m.group(1)

    if v.startswith("+"):
        # invite link token (cannot be used as chat_id)
        return None

    if v.startswith("@"):  # @channel
        v = v[1:]

    v = v.strip()
    if not re.fullmatch(r"[A-Za-z0-9_]{4,}", v):
        return None
    return f"@{v}"


def _candidate_bot_username(candidate: models.User) -> str | None:
    v = (getattr(candidate, "bot_name", None) or "").strip()
    if v.startswith("@"):  # tolerate @BotUsername
        v = v[1:]
    v = v.split()[0] if v else ""
    if not v:
        return None
    if not re.fullmatch(r"[A-Za-z0-9_]{4,}", v):
        return None
    return v


def _notify_question_answer_published(
    *,
    candidate: models.User,
    submission: models.BotSubmission,
):
    """Best-effort: notify candidate socials (group + channel) when a question gets answered.

    This must never fail the API request.
    """
    try:
        token = (getattr(candidate, "bot_token", None) or "").strip()
        if not token:
            return

        socials = getattr(candidate, "socials", None) or {}
        if not isinstance(socials, dict):
            socials = {}

        # Prefer explicit numeric chat ids if present.
        group_chat_id = socials.get("telegram_group_chat_id") or socials.get("telegramGroupChatId")
        channel_chat_id = socials.get("telegram_channel_chat_id") or socials.get("telegramChannelChatId")

        group_raw = socials.get("telegram_group") or socials.get("telegramGroup")
        channel_raw = socials.get("telegram_channel") or socials.get("telegramChannel")
        targets = [
            _extract_telegram_chat_target(str(group_chat_id) if group_chat_id is not None else None)
            or _extract_telegram_chat_target(str(group_raw) if group_raw is not None else None),
            _extract_telegram_chat_target(str(channel_chat_id) if channel_chat_id is not None else None)
            or _extract_telegram_chat_target(str(channel_raw) if channel_raw is not None else None),
        ]
        targets = [t for t in targets if t is not None]
        # De-duplicate while preserving order
        seen: set[str] = set()
        uniq: list[str | int] = []
        for t in targets:
            key = str(t)
            if key in seen:
                continue
            seen.add(key)
            uniq.append(t)

        if not uniq:
            return

        bot_username = _candidate_bot_username(candidate)
        deep_link = (
            f"https://t.me/{bot_username}?start=question_{int(submission.id)}" if bot_username else None
        )

        topic = (getattr(submission, "topic", None) or "").strip()
        topic_line = f"\n🗂 دسته: {topic}" if topic else ""
        link_line = f"\n\n🔗 مشاهده در بات: {deep_link}" if deep_link else ""

        text = (
            "✅ پاسخ جدید به سؤال مردمی منتشر شد."
            f"\n🔖 کد سؤال: {int(submission.id)}"
            f"{topic_line}"
            f"{link_line}"
        )

        url = f"https://api.telegram.org/bot{token}/sendMessage"
        with httpx.Client(timeout=10.0) as client:
            for chat_id in uniq:
                try:
                    resp = client.post(
                        url,
                        json={
                            "chat_id": chat_id,
                            "text": text,
                            "disable_web_page_preview": True,
                        },
                    )
                    # If we used @username and Telegram returns a numeric chat.id, persist it.
                    try:
                        if resp.is_success and isinstance(chat_id, str) and str(chat_id).startswith("@"):  # @username
                            payload = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else None
                            chat_obj = (payload or {}).get("result", {}).get("chat", {})
                            numeric_id = chat_obj.get("id")
                            if isinstance(numeric_id, int):
                                # Store channel numeric id if this was the channel username.
                                if str(chat_id) == str(_extract_telegram_chat_target(str(channel_raw) if channel_raw is not None else None)):
                                    db2 = database.SessionLocal()
                                    try:
                                        u2 = db2.query(models.User).filter(models.User.id == int(candidate.id)).first()
                                        if u2:
                                            base2 = u2.socials if isinstance(u2.socials, dict) else {}
                                            next2 = dict(base2)
                                            if next2.get("telegram_channel_chat_id") != numeric_id:
                                                next2["telegram_channel_chat_id"] = numeric_id
                                            if next2.get("telegramChannelChatId") != numeric_id:
                                                next2["telegramChannelChatId"] = numeric_id
                                            u2.socials = next2
                                            db2.add(u2)
                                            db2.commit()
                                    finally:
                                        db2.close()
                    except Exception:
                        logger.debug("Failed to persist numeric chat id from Telegram response")
                    if not resp.is_success:
                        logger.warning(
                            "Question notify failed: status=%s body=%s",
                            resp.status_code,
                            (resp.text or "")[:500],
                        )
                except Exception:
                    logger.exception("Question notify exception")
    except Exception:
        logger.exception("Question notify wrapper exception")

# ساخت جداول دیتابیس
models.Base.metadata.create_all(bind=database.engine)

@asynccontextmanager
async def lifespan(app: FastAPI):
    print("Application started")
    yield
    print("Application shutdown")

app = FastAPI(
    title="Election Manager",
    version="1.0.0",
    lifespan=lifespan
)


@app.middleware("http")
async def monitoring_error_logging_middleware(request, call_next):
    """Minimal technical error logging (MVP).

    Goal: make 5xx and unhandled exceptions visible to super-admin monitoring.
    """
    try:
        response = await call_next(request)
        if getattr(response, "status_code", 200) >= 500:
            try:
                db = database.SessionLocal()
                db.add(
                    models.TechnicalErrorLog(
                        service_name="api",
                        error_type="HTTP_5XX",
                        error_message=f"{request.method} {request.url.path} -> {response.status_code}",
                        telegram_user_id=None,
                        candidate_id=None,
                        state=None,
                    )
                )
                db.commit()
            except Exception:
                try:
                    db.rollback()
                except Exception:
                    pass
            finally:
                try:
                    db.close()
                except Exception:
                    pass
        return response
    except Exception as e:
        try:
            db = database.SessionLocal()
            db.add(
                models.TechnicalErrorLog(
                    service_name="api",
                    error_type=e.__class__.__name__,
                    error_message=str(e)[:8000],
                    telegram_user_id=None,
                    candidate_id=None,
                    state=None,
                )
            )
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
        finally:
            try:
                db.close()
            except Exception:
                pass
        raise

# Mount static files for uploads
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_origins=(
        [o.strip() for o in (os.getenv("CORS_ALLOW_ORIGINS") or "").split(",") if o.strip()]
        if APP_ENV in {"production", "prod"} and (os.getenv("CORS_ALLOW_ORIGINS") or "").strip()
        else [
            "http://localhost:5173",
            "http://127.0.0.1:5173",
            "http://localhost:5174",
            "http://127.0.0.1:5174",
            "http://localhost:3000",
            "http://127.0.0.1:3000",
            "http://localhost:3001",
            "http://127.0.0.1:3001",
            "http://localhost:3002",
            "http://127.0.0.1:3002",
            "http://localhost:5555",
            "http://127.0.0.1:5555",
        ]
    ),
    # Dev convenience: Vite may auto-select another free 5xxx port.
    allow_origin_regex=None if APP_ENV in {"production", "prod"} else r"^http://(localhost|127\\.0\\.0\\.1):5\\d{3}$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "no-referrer")
    response.headers.setdefault("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
    return response

@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...),
    candidate_name: str | None = Form(None),
    current_user: models.User = Depends(auth.get_current_user),
    request: Request = None,
):
    def _safe_part(value: str) -> str:
        v = (value or "").strip()
        v = re.sub(r"\s+", "_", v)
        # Keep unicode letters/digits/underscore plus .-_ ; drop everything else.
        v = re.sub(r"[^\w\-\.]+", "", v, flags=re.UNICODE)
        v = v.strip("._-")
        return (v[:80] or "file")

    original_name = (file.filename or "").strip()
    original_ext = os.path.splitext(original_name)[1].lower()
    if not original_ext or len(original_ext) > 10:
        guessed = mimetypes.guess_extension(file.content_type or "") or ""
        original_ext = guessed if guessed.startswith(".") else ""
    ext = original_ext or ".bin"

    denied_exts = {
        ".html",
        ".htm",
        ".js",
        ".mjs",
        ".cjs",
        ".svg",
        ".xml",
        ".php",
        ".py",
        ".sh",
        ".bat",
        ".ps1",
        ".exe",
        ".dll",
        ".msi",
        ".jar",
    }
    if ext in denied_exts:
        raise HTTPException(status_code=422, detail={"message": "این نوع فایل مجاز نیست."})

    prefix = "معرفی-صوتی"
    who = _safe_part(candidate_name or "")
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    suffix = uuid.uuid4().hex[:6]
    base = f"{prefix}-{who}" if who else prefix
    filename = f"{base}-{ts}-{suffix}{ext}"

    MAX_BYTES = 10 * 1024 * 1024  # 10MB default (override via UPLOAD_MAX_BYTES)
    try:
        env_max = int((os.getenv("UPLOAD_MAX_BYTES") or "").strip() or "0")
        if env_max > 0:
            MAX_BYTES = env_max
    except Exception:
        pass

    file_location = os.path.join(UPLOAD_DIR, filename)
    written = 0
    try:
        with open(file_location, "wb+") as buffer:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                written += len(chunk)
                if written > MAX_BYTES:
                    raise HTTPException(status_code=413, detail={"message": "حجم فایل بیش از حد مجاز است."})
                buffer.write(chunk)
    except HTTPException:
        try:
            if os.path.exists(file_location):
                os.remove(file_location)
        except Exception:
            pass
        raise
    finally:
        try:
            await file.close()
        except Exception:
            pass

    base = ""
    try:
        if request is not None and request.base_url:
            base = str(request.base_url).rstrip("/")
    except Exception:
        base = ""
    url = f"{base}/uploads/{filename}" if base else f"/uploads/{filename}"
    return {"url": url, "filename": filename}


@app.post("/api/upload/voice-intro")
async def upload_voice_intro(
    file: UploadFile = File(...),
    candidate_name: str | None = Form(None),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Upload candidate voice introduction (MVP).

    Hard constraints:
    - Single file per candidate (enforced by overwriting candidate.voice_url via update endpoint)
    - Max size: 2MB
    - Format: mp3 or ogg
    """
    MAX_BYTES = 2 * 1024 * 1024
    allowed_exts = {".mp3", ".ogg"}
    allowed_ct_prefixes = ("audio/",)
    allowed_ct_exact = {"application/ogg"}

    original_name = (file.filename or "").strip()
    ext = os.path.splitext(original_name)[1].lower()
    content_type = (file.content_type or "").strip().lower()

    if ext not in allowed_exts:
        raise HTTPException(
            status_code=422,
            detail={"message": "فرمت فایل صوتی باید mp3 یا ogg باشد."},
        )

    if content_type and not (content_type.startswith(allowed_ct_prefixes) or content_type in allowed_ct_exact):
        raise HTTPException(
            status_code=422,
            detail={"message": "نوع فایل معتبر نیست. فقط فایل صوتی مجاز است."},
        )

    # Reuse safe naming logic
    def _safe_part(value: str) -> str:
        v = (value or "").strip()
        v = re.sub(r"\s+", "_", v)
        v = re.sub(r"[^\w\-\.]+", "", v, flags=re.UNICODE)
        v = v.strip("._-")
        return (v[:80] or "file")

    prefix = "معرفی-صوتی"
    who = _safe_part(candidate_name or "")
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    suffix = uuid.uuid4().hex[:6]
    base = f"{prefix}-{who}" if who else prefix
    filename = f"{base}-{ts}-{suffix}{ext}"
    file_location = os.path.join(UPLOAD_DIR, filename)

    written = 0
    try:
        with open(file_location, "wb+") as buffer:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                written += len(chunk)
                if written > MAX_BYTES:
                    raise HTTPException(
                        status_code=413,
                        detail={"message": "حجم فایل صوتی باید حداکثر ۲ مگابایت باشد."},
                    )
                buffer.write(chunk)
    except HTTPException:
        try:
            if os.path.exists(file_location):
                os.remove(file_location)
        except Exception:
            pass
        raise
    finally:
        try:
            await file.close()
        except Exception:
            pass

    return {"url": f"http://127.0.0.1:8000/uploads/{filename}", "filename": filename}


def _upload_file_path_from_localhost_url(url: str | None) -> str | None:
    if not url or not isinstance(url, str):
        return None
    url = url.strip()
    if not url:
        return None
    prefixes = (
        "http://localhost:8000/uploads/",
        "http://127.0.0.1:8000/uploads/",
    )
    if not url.startswith(prefixes):
        return None
    filename = url.split("/uploads/", 1)[-1]
    filename = filename.split("?", 1)[0].split("#", 1)[0]
    filename = filename.replace("..", "").lstrip("/\\")
    local_path = os.path.join(os.path.dirname(__file__), UPLOAD_DIR, filename)
    return local_path if os.path.exists(local_path) else None


def _telegram_post_json(token: str, method: str, payload: dict) -> tuple[bool, str | None]:
    url = f"https://api.telegram.org/bot{token}/{method}"
    try:
        with httpx.Client(timeout=20) as client:
            res = client.post(url, json=payload)
        data = res.json() if res.content else {}
        if res.status_code >= 400 or not data.get("ok"):
            desc = data.get("description") or f"HTTP {res.status_code}"
            return False, _humanize_telegram_error(str(desc))
        return True, None
    except Exception as e:
        return False, _humanize_telegram_error(str(e))


def _telegram_post_file(
    token: str,
    method: str,
    *,
    field_name: str,
    filename: str,
    content: bytes,
    content_type: str,
) -> tuple[bool, str | None]:
    url = f"https://api.telegram.org/bot{token}/{method}"
    try:
        with httpx.Client(timeout=30) as client:
            res = client.post(url, files={field_name: (filename, content, content_type)})
        data = res.json() if res.content else {}
        if res.status_code >= 400 or not data.get("ok"):
            desc = data.get("description") or f"HTTP {res.status_code}"
            return False, _humanize_telegram_error(str(desc))
        return True, None
    except Exception as e:
        return False, _humanize_telegram_error(str(e))


_BOT_TOKEN_RE = re.compile(r"^\d+:[A-Za-z0-9_-]{20,}$")


def _looks_like_telegram_bot_token(token: str) -> bool:
    token = (token or "").strip()
    return bool(_BOT_TOKEN_RE.match(token))


def _telegram_get_me(token: str) -> tuple[bool, str | None]:
    """Preflight check to distinguish invalid token vs method unsupported."""
    url = f"https://api.telegram.org/bot{token}/getMe"
    try:
        with httpx.Client(timeout=20) as client:
            res = client.get(url)
        data = res.json() if res.content else {}
        if res.status_code >= 400 or not data.get("ok"):
            desc = data.get("description") or f"HTTP {res.status_code}"
            if res.status_code == 404 and str(desc).strip() == "Not Found":
                return False, "توکن بات نامعتبر است (Telegram: Not Found). لطفاً Bot Token را دقیق از BotFather کپی کنید."
            return False, _humanize_telegram_error(str(desc))
        return True, None
    except Exception as e:
        return False, _humanize_telegram_error(str(e))


def _telegram_get_json(token: str, method: str) -> tuple[bool, dict | None, str | None]:
    url = f"https://api.telegram.org/bot{token}/{method}"
    try:
        with httpx.Client(timeout=20) as client:
            res = client.get(url)
        data = res.json() if res.content else {}
        if res.status_code >= 400 or not data.get("ok"):
            desc = data.get("description") or f"HTTP {res.status_code}"
            if res.status_code == 404 and str(desc).strip() == "Not Found":
                return False, None, "پاسخ تلگرام: Not Found (احتمالاً این متد برای این نسخه/اکانت پشتیبانی نمی‌شود)."
            return False, None, _humanize_telegram_error(str(desc))
        return True, data.get("result") if isinstance(data, dict) else None, None
    except Exception as e:
        return False, None, _humanize_telegram_error(str(e))


def _humanize_telegram_error(err: str) -> str:
    msg = (err or "").strip()
    if not msg:
        return "خطای نامشخص هنگام ارتباط با تلگرام"

    lower = msg.lower()

    if msg.strip() == "Not Found" or "telegram: not found" in lower:
        return "پاسخ تلگرام: Not Found (احتمالاً متد پشتیبانی نمی‌شود یا دسترسی شما به Bot API محدود است)."

    if "unexpected_eof_while_reading" in lower or "eof occurred in violation of protocol" in lower:
        return "اتصال امن (SSL) به تلگرام قطع شد. معمولاً به خاطر اینترنت/فیلترینگ/پروکسی یا آنتی‌ویروس است."

    if "certificate_verify_failed" in lower:
        return "اعتبارسنجی گواهی SSL تلگرام ناموفق بود (CERTIFICATE_VERIFY_FAILED). تاریخ/ساعت سیستم و تنظیمات پروکسی را بررسی کنید."

    if "timed out" in lower or "timeout" in lower:
        return "ارتباط با تلگرام timeout شد. اینترنت/فیلترشکن/پروکسی را بررسی کنید."

    return msg


def _guess_image_content_type(filename: str) -> str:
    ct, _ = mimetypes.guess_type(filename)
    return ct or "application/octet-stream"


def apply_telegram_profile_for_candidate(candidate: models.User) -> dict:
    """Apply Telegram profile settings using the saved candidate fields.

    Best-effort: returns a status payload with per-field results.
    """
    result: dict = {"ok": True, "applied": {}, "errors": {}, "notes": {}, "requested": {}, "telegram": {}}

    token = (candidate.bot_token or "").strip()
    if not token:
        result["ok"] = False
        result["errors"]["token"] = "توکن بات تنظیم نشده است."
        return result

    if not _looks_like_telegram_bot_token(token):
        result["ok"] = False
        result["errors"]["token"] = "فرمت توکن بات نامعتبر است. نمونه صحیح: 123456789:AA... (از BotFather کپی کنید)."
        return result

    ok, err = _telegram_get_me(token)
    result["applied"]["getMe"] = ok
    if not ok:
        result["ok"] = False
        result["errors"]["token"] = err or "بررسی توکن بات در تلگرام ناموفق بود"
        return result

    # 1) Name (Telegram display name)
    if candidate.bot_name and isinstance(candidate.bot_name, str) and candidate.bot_name.strip():
        result["requested"]["name"] = candidate.bot_name.strip()
        ok, err = _telegram_post_json(token, "setMyName", {"name": candidate.bot_name.strip()})
        result["applied"]["name"] = ok
        if not ok:
            result["ok"] = False
            result["errors"]["name"] = err

    # 2) Description / short description from bot_config
    bot_config = candidate.bot_config or {}
    telegram_profile: dict = {}
    if isinstance(bot_config, dict):
        telegram_profile = bot_config.get("telegram_profile") or bot_config.get("telegramProfile") or {}

    if isinstance(telegram_profile, dict):
        desc = telegram_profile.get("description")
        if isinstance(desc, str):
            result["requested"]["description"] = desc
            ok, err = _telegram_post_json(token, "setMyDescription", {"description": desc})
            result["applied"]["description"] = ok
            if not ok:
                result["ok"] = False
                result["errors"]["description"] = err

        short_desc = telegram_profile.get("short_description") or telegram_profile.get("shortDescription")
        if isinstance(short_desc, str):
            result["requested"]["short_description"] = short_desc
            ok, err = _telegram_post_json(token, "setMyShortDescription", {"short_description": short_desc})
            result["applied"]["short_description"] = ok
            if not ok:
                result["ok"] = False
                result["errors"]["short_description"] = err

    # 3) Profile photo: Telegram Bot API does not support changing bot avatar via HTTP API.
    # Keep image_url for in-chat photo sending and UI preview.
    image_url = candidate.image_url
    if isinstance(image_url, str) and image_url.strip():
        result["applied"]["photo"] = False
        result["notes"]["photo"] = (
            "تغییر عکس پروفایل بات از طریق Telegram Bot API پشتیبانی نمی‌شود. "
            "برای تغییر آواتار، باید دستی از طریق BotFather انجام دهید (مثلاً دستور /setuserpic)."
        )

    # 4) Read back current Telegram profile values for verification
    ok_name, res_name, _ = _telegram_get_json(token, "getMyName")
    if ok_name and isinstance(res_name, dict) and isinstance(res_name.get("name"), str):
        result["telegram"]["name"] = res_name.get("name")

    ok_desc, res_desc, _ = _telegram_get_json(token, "getMyDescription")
    if ok_desc and isinstance(res_desc, dict) and isinstance(res_desc.get("description"), str):
        result["telegram"]["description"] = res_desc.get("description")

    ok_sdesc, res_sdesc, _ = _telegram_get_json(token, "getMyShortDescription")
    if ok_sdesc and isinstance(res_sdesc, dict) and isinstance(res_sdesc.get("short_description"), str):
        result["telegram"]["short_description"] = res_sdesc.get("short_description")

    return result

# ============================================================================
# ✅ HELPERS برای خطاهای Duplicate Field
# ============================================================================

FIELD_LABELS = {
    "username": "نام کاربری",
    "phone": "شماره تماس",
    "email": "ایمیل",
    "bot_name": "نام بات",
    "bot_token": "توکن بات",
}

def raise_duplicate_field(field: str):
    """خطای 400 برای فیلد تکراری"""
    label = FIELD_LABELS.get(field, field)
    raise HTTPException(
        status_code=400,
        detail={
            "code": "DUPLICATE_FIELD",
            "field": field,
            "label": label,
            "message": f"«{label}» تکراری است. لطفاً مقدار دیگری وارد کنید.",
        },
    )

def parse_integrity_error_field(e: IntegrityError) -> Optional[str]:
    """SQLite error از IntegrityError فیلد را استخراج کند"""
    msg = str(getattr(e, "orig", e))
    msg_lower = msg.lower()

    if "unique constraint failed:" in msg_lower:
        tail = msg.split("UNIQUE constraint failed:", 1)[-1].strip()
        first_part = tail.split(",")[0].strip()
        field = first_part.split(".")[-1].strip() if "." in first_part else first_part
        return field if field else None

    return None

def raise_from_integrity_error(e: IntegrityError):
    """IntegrityError را به پیام کاربرپسند تبدیل کن"""
    field = parse_integrity_error_field(e)
    if field:
        raise_duplicate_field(field)

    raise HTTPException(
        status_code=400,
        detail={
            "code": "DUPLICATE_FIELD",
            "message": "یکی از فیلدها تکراری است. لطفاً مقادیر را تغییر دهید.",
        },
    )

# ============================================================================
# REQUEST/RESPONSE SCHEMAS
# ============================================================================

class LoginRequest(BaseModel):
    username: str
    password: str

class PasswordResetRequest(BaseModel):
    password: str

class TokenResponse(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str

class RefreshRequest(BaseModel):
    refresh_token: str

class RegisterRequest(BaseModel):
    username: str
    password: str
    email: str
    full_name: Optional[str] = None

class AssignPlanRequest(BaseModel):
    plan_id: int
    duration_days: int = 30

# ============================================================================
# AUTH ENDPOINTS
# ============================================================================

@app.post("/api/auth/register", response_model=TokenResponse)
def register(request: RegisterRequest, req: Request, db: Session = Depends(database.get_db)):
    """ثبت نام کاربر جدید.

    امنیت: این endpoint نباید نقش ADMIN بسازد. ثبت‌نام عمومی فقط کاربر CANDIDATE می‌سازد.
    """
    _rate_limit(req, key="auth:register", limit=10, window_seconds=60)

    username = (request.username or "").strip()
    email = (request.email or "").strip()
    password = (request.password or "").strip()
    full_name = (request.full_name or "").strip() or None

    if not username or not password or not email:
        raise HTTPException(status_code=422, detail="نام کاربری، ایمیل و رمز عبور الزامی است")
    if len(password) < 8:
        raise HTTPException(status_code=422, detail="رمز عبور باید حداقل ۸ کاراکتر باشد")

    existing_user = db.query(models.User).filter(models.User.username == username).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="نام کاربری تکراری است")

    existing_email = db.query(models.User).filter(models.User.email == email).first()
    if existing_email:
        raise HTTPException(status_code=400, detail="ایمیل تکراری است")

    new_user = models.User(
        username=username,
        email=email,
        full_name=full_name,
        hashed_password=auth.get_password_hash(password),
        role="CANDIDATE",
        is_active=True,
    )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    return auth.create_tokens(new_user.username)

@app.post("/api/auth/login", response_model=TokenResponse)
def login(request: LoginRequest, req: Request, db: Session = Depends(database.get_db)):
    _rate_limit(req, key="auth:login", limit=20, window_seconds=60)
    user = auth.authenticate_user(db, request.username, request.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="نام کاربری یا رمز عبور اشتباه است",
        )
    tokens = auth.create_tokens(user.username)
    return tokens


@app.post("/api/auth/refresh", response_model=TokenResponse)
def refresh_access_token(request: RefreshRequest, req: Request, db: Session = Depends(database.get_db)):
    """Issue a new access token using a valid refresh token."""
    _rate_limit(req, key="auth:refresh", limit=30, window_seconds=60)
    username = auth.decode_refresh_token(request.refresh_token)
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user or not user.is_active:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Could not validate credentials")
    return auth.create_tokens(username)

@app.get("/api/auth/me", response_model=schemas.User)
def get_current_user(
    current_user: models.User = Depends(auth.get_current_user),
):
    # Ensure we never serialize sensitive columns like hashed_password.
    return current_user

# ============================================================================
# CANDIDATES ENDPOINTS (Now operating on Users with role=CANDIDATE)
# ============================================================================

@app.get("/api/candidates", response_model=List[schemas.Candidate])
def get_candidates(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    # Security: this endpoint previously leaked bot_token/bot_config publicly.
    # Now it requires auth.
    if current_user.role == "ADMIN":
        return (
            db.query(models.User)
            .filter(models.User.role == "CANDIDATE")
            .order_by(models.User.id.desc())
            .all()
        )

    if current_user.role == "CANDIDATE":
        return [current_user]

    raise HTTPException(status_code=403, detail="Access denied")

@app.get("/api/candidates/{candidate_id}", response_model=schemas.Candidate)
def get_candidate(
    candidate_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    candidate = (
        db.query(models.User)
        .filter(models.User.id == candidate_id, models.User.role == "CANDIDATE")
        .first()
    )
    if not candidate:
        raise HTTPException(status_code=404, detail="کاندید یافت نشد")

    if current_user.role != "ADMIN" and candidate.id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    return candidate

@app.post("/api/candidates", response_model=schemas.Candidate)
def create_candidate(
    candidate_data: schemas.CandidateCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    data = candidate_data.model_dump(exclude_unset=True)

    try:
        username = (data.get("username") or "").strip()
        password = (data.get("password") or "").strip()
        phone = data.get("phone")
        full_name = (data.get("name") or "").strip()
        bot_name = (data.get("bot_name") or "").strip()
        bot_token = (data.get("bot_token") or "").strip()

        if isinstance(phone, str):
            phone = phone.strip() or None
        if isinstance(bot_name, str):
            bot_name = bot_name.strip() or None
        if isinstance(bot_token, str):
            bot_token = bot_token.strip() or None

        if not username or not password:
            raise HTTPException(
                status_code=422,
                detail={"message": "نام کاربری و رمز عبور برای ایجاد نماینده الزامی است"},
            )

        # ✅ تاریخ شمسی
        now_jalali = jdatetime.datetime.now()
        created_at_jalali = now_jalali.strftime("%Y/%m/%d %H:%M:%S")

        new_user = models.User(
            username=username,
            phone=phone,
            full_name=full_name,
            hashed_password=auth.get_password_hash(password),
            role="CANDIDATE",
            is_active=True,
            bot_name=bot_name,
            bot_token=bot_token,
            city=data.get("city") or None,
            province=data.get("province") or None,
            constituency=(data.get("constituency") or None),
            slogan=data.get("slogan"),
            bio=data.get("bio"),
            image_url=data.get("image_url"),
            resume=data.get("resume"),
            ideas=data.get("ideas"),
            address=data.get("address"),
            voice_url=data.get("voice_url"),
            socials=data.get("socials"),
            bot_config=data.get("bot_config"),
            created_at_jalali=created_at_jalali,
        )
        
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        return new_user

    except IntegrityError as e:
        db.rollback()
        raise_from_integrity_error(e)
    except HTTPException:
        db.rollback()
        raise

@app.put("/api/candidates/{candidate_id}", response_model=schemas.Candidate)
def update_candidate(
    candidate_id: int,
    candidate_data: schemas.CandidateUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    candidate = db.query(models.User).filter(models.User.id == candidate_id, models.User.role == "CANDIDATE").first()
    if not candidate:
        raise HTTPException(status_code=404, detail="کاندید یافت نشد")

    # Check permissions: Admin or the candidate themselves
    if current_user.role != "ADMIN":
        if candidate.id != current_user.id:
            raise HTTPException(status_code=403, detail="شما اجازه ویرایش این کاندید را ندارید")

    data = candidate_data.model_dump(exclude_unset=True) if hasattr(candidate_data, "model_dump") else candidate_data.dict()

    try:
        previous_voice_url = candidate.voice_url

        # Handle password update separately
        if "password" in data and data["password"]:
            candidate.hashed_password = auth.get_password_hash(data["password"])
            del data["password"]

        # Map 'name' to 'full_name'
        if "name" in data:
            candidate.full_name = data["name"]
            del data["name"]

        for key, value in data.items():
            if not hasattr(candidate, key):
                continue

            # Allow explicitly clearing some nullable fields by sending null.
            if value is None:
                if key in {"voice_url"}:
                    setattr(candidate, key, None)
                continue

            if key in {"city", "province", "constituency"} and isinstance(value, str):
                value = value.strip() or None

            setattr(candidate, key, value)

        # If voice has been cleared or replaced, best-effort delete the previous uploaded file.
        if "voice_url" in data:
            new_voice_url = candidate.voice_url
            if previous_voice_url and previous_voice_url != new_voice_url:
                try:
                    old_path = _upload_file_path_from_localhost_url(previous_voice_url)
                    if old_path and os.path.exists(old_path):
                        os.remove(old_path)
                except Exception:
                    # best-effort cleanup only
                    pass

        db.commit()
        db.refresh(candidate)
        return candidate

    except IntegrityError as e:
        db.rollback()
        raise_from_integrity_error(e)
    except HTTPException:
        db.rollback()
        raise


@app.post("/api/candidates/{candidate_id}/apply-telegram-profile")
def apply_telegram_profile(
    candidate_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Apply Telegram bot profile settings (photo/name/description) for a candidate.

    Returns a status payload with ok/applied/errors. Does not fail the request on partial failures
    so the UI can present granular errors.
    """
    candidate = (
        db.query(models.User)
        .filter(models.User.id == candidate_id, models.User.role == "CANDIDATE")
        .first()
    )
    if not candidate:
        raise HTTPException(status_code=404, detail="کاندید یافت نشد")

    if current_user.role != "ADMIN" and candidate.id != current_user.id:
        raise HTTPException(status_code=403, detail="شما اجازه این عملیات را ندارید")

    status_payload = apply_telegram_profile_for_candidate(candidate)
    return status_payload

@app.delete("/api/candidates/{candidate_id}")
def delete_candidate(
    candidate_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    candidate = db.query(models.User).filter(models.User.id == candidate_id, models.User.role == "CANDIDATE").first()
    if not candidate:
        raise HTTPException(status_code=404, detail="کاندید یافت نشد")

    db.delete(candidate)
    db.commit()
    return {"detail": "کاندید حذف شد"}

@app.post("/api/candidates/{candidate_id}/reset-password")
def reset_candidate_password(
    candidate_id: int,
    body: PasswordResetRequest,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    candidate = db.query(models.User).filter(models.User.id == candidate_id, models.User.role == "CANDIDATE").first()
    if not candidate:
        raise HTTPException(status_code=404, detail="کاندید یافت نشد")

    candidate.hashed_password = auth.get_password_hash(body.password)
    db.add(candidate)
    db.commit()
    return {"detail": "رمز عبور با موفقیت تغییر کرد"}

@app.post("/api/candidates/{candidate_id}/assign-plan")
def assign_plan_to_candidate(
    candidate_id: int,
    request: AssignPlanRequest,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    candidate = db.query(models.User).filter(models.User.id == candidate_id, models.User.role == "CANDIDATE").first()
    if not candidate:
        raise HTTPException(status_code=404, detail="کاندید یافت نشد")
    
    plan = db.query(models.Plan).filter(models.Plan.id == request.plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="پلن یافت نشد")

    now = datetime.utcnow()
    candidate.active_plan_id = plan.id
    candidate.plan_start_date = now
    candidate.plan_expires_at = now + timedelta(days=request.duration_days)
    
    db.commit()
    return {"message": "پلن با موفقیت فعال شد"}

# ============================================================================
# PLANS ENDPOINTS
# ============================================================================

@app.get("/api/plans")
def get_plans(db: Session = Depends(database.get_db)):
    return db.query(models.Plan).order_by(models.Plan.id.desc()).all()

@app.post("/api/plans", response_model=schemas.Plan)
def create_plan(
    plan_data: schemas.PlanCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    data = plan_data.model_dump(exclude_unset=True)
    
    now_jalali = jdatetime.datetime.now()
    created_at_jalali = now_jalali.strftime("%Y/%m/%d %H:%M:%S")
    
    new_plan = models.Plan(**data, created_at_jalali=created_at_jalali)
    db.add(new_plan)
    db.commit()
    db.refresh(new_plan)
    return new_plan

@app.put("/api/plans/{plan_id}", response_model=schemas.Plan)
def update_plan(
    plan_id: int,
    plan_data: schemas.PlanUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    plan = db.query(models.Plan).filter(models.Plan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="پلن یافت نشد")

    data = plan_data.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(plan, key, value)

    db.commit()
    db.refresh(plan)
    return plan

@app.delete("/api/plans/{plan_id}")
def delete_plan(
    plan_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    plan = db.query(models.Plan).filter(models.Plan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="پلن یافت نشد")

    db.delete(plan)
    db.commit()
    return {"detail": "پلن حذف شد"}

# ============================================================================
# TICKETS ENDPOINTS
# ============================================================================

@app.get("/api/tickets", response_model=List[schemas.Ticket])
def get_tickets(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    # Security: never expose all tickets publicly.
    q = db.query(models.Ticket).order_by(models.Ticket.id.desc())
    if current_user.role == "ADMIN":
        return q.all()
    return q.filter(models.Ticket.user_id == current_user.id).all()

@app.post("/api/tickets", response_model=schemas.Ticket)
def create_ticket(
    ticket_data: schemas.TicketCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    # If user is candidate, they are creating ticket for themselves
    # If user is admin, they might be creating for someone else (not implemented yet, assuming self)
    
    if current_user.role == "CANDIDATE":
        user_id = current_user.id
    else:
        # Admin creating ticket? For now let's say admin creates for themselves or we need a field
        user_id = current_user.id

    new_ticket = models.Ticket(
        user_id=user_id,
        subject=ticket_data.subject,
        status="OPEN"
    )
    db.add(new_ticket)
    db.flush() # Get ID

    # Add initial message
    initial_msg = models.TicketMessage(
        ticket_id=new_ticket.id,
        sender_role=current_user.role,
        text=ticket_data.message
    )
    db.add(initial_msg)
    db.commit()
    db.refresh(new_ticket)
    return new_ticket

@app.post("/api/tickets/{ticket_id}/messages", response_model=schemas.TicketMessage)
def add_ticket_message(
    ticket_id: int,
    message_data: schemas.TicketMessageCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    ticket = db.query(models.Ticket).filter(models.Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    # Authorization: admin can access all; candidate only own tickets.
    if current_user.role != "ADMIN" and ticket.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    new_msg = models.TicketMessage(
        ticket_id=ticket_id,
        # Security: do not trust client-provided sender_role.
        sender_role=current_user.role,
        text=message_data.text,
        attachment_url=message_data.attachment_url,
        attachment_type=message_data.attachment_type
    )
    
    ticket.updated_at = datetime.utcnow()
    if current_user.role == "ADMIN":
        ticket.status = "ANSWERED"
    else:
        ticket.status = "OPEN"

    db.add(new_msg)
    db.commit()
    db.refresh(new_msg)
    return new_msg

@app.put("/api/tickets/{ticket_id}/status")
def update_ticket_status(
    ticket_id: int,
    update: schemas.TicketUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user)
):
    ticket = db.query(models.Ticket).filter(models.Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    ticket.status = update.status
    db.commit()
    return {"message": "Status updated"}

# ============================================================================
# FEEDBACK (BOT SUBMISSIONS) ENDPOINTS (MVP)
# ============================================================================

@app.get("/api/candidates/me/feedback", response_model=List[schemas.FeedbackSubmission])
def get_my_feedback_submissions(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.role != "CANDIDATE":
        raise HTTPException(status_code=403, detail="Access denied")

    return (
        db.query(models.BotSubmission)
        .filter(
            models.BotSubmission.candidate_id == current_user.id,
            models.BotSubmission.type == "FEEDBACK",
        )
        .order_by(models.BotSubmission.id.desc())
        .all()
    )


# ============================================================================
# ADMIN: BOT BUILD REQUESTS (MVP)
# ============================================================================


@app.get("/api/admin/bot-requests", response_model=List[schemas.BotRequestSubmission])
def admin_list_bot_requests(
    status: Optional[str] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    q = (
        db.query(models.BotSubmission)
        .filter(models.BotSubmission.type == "BOT_REQUEST")
        .order_by(models.BotSubmission.id.desc())
    )
    if status:
        q = q.filter(models.BotSubmission.status == status)
    return q.all()


@app.put("/api/admin/bot-requests/{submission_id}", response_model=schemas.BotRequestSubmission)
def admin_update_bot_request(
    submission_id: int,
    body: schemas.BotRequestUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    item = (
        db.query(models.BotSubmission)
        .filter(models.BotSubmission.id == submission_id, models.BotSubmission.type == "BOT_REQUEST")
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="درخواست یافت نشد")

    item.status = (body.status or "").strip() or item.status
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.get("/api/admin/dashboard-stats", response_model=schemas.AdminDashboardStats)
def admin_dashboard_stats(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    active_bots = (
        db.query(func.count(models.User.id))
        .filter(
            models.User.role == "CANDIDATE",
            models.User.is_active == True,  # noqa: E712
            models.User.bot_token.isnot(None),
        )
        .scalar()
        or 0
    )

    total_questions = (
        db.query(func.count(models.BotSubmission.id))
        .filter(models.BotSubmission.type == "QUESTION")
        .scalar()
        or 0
    )
    total_feedback = (
        db.query(func.count(models.BotSubmission.id))
        .filter(models.BotSubmission.type == "FEEDBACK")
        .scalar()
        or 0
    )
    total_bot_requests = (
        db.query(func.count(models.BotSubmission.id))
        .filter(models.BotSubmission.type == "BOT_REQUEST")
        .scalar()
        or 0
    )

    return schemas.AdminDashboardStats(
        active_bots=active_bots,
        total_questions=total_questions,
        total_feedback=total_feedback,
        total_bot_requests=total_bot_requests,
    )


@app.get("/api/admin/candidate-stats", response_model=List[schemas.AdminCandidateStats])
def admin_candidate_stats(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    # Aggregate counts from BotSubmission (single table holds QUESTION/FEEDBACK)
    answered_cond = (
        (models.BotSubmission.type == "QUESTION")
        & (func.upper(func.coalesce(models.BotSubmission.status, "")) == "ANSWERED")
    )

    rows = (
        db.query(
            models.BotSubmission.candidate_id.label("candidate_id"),
            func.sum(case((models.BotSubmission.type == "QUESTION", 1), else_=0)).label("total_questions"),
            func.sum(case((models.BotSubmission.type == "FEEDBACK", 1), else_=0)).label("total_feedback"),
            func.sum(case((answered_cond, 1), else_=0)).label("answered_questions"),
        )
        .group_by(models.BotSubmission.candidate_id)
        .all()
    )

    # Ensure ints in response (SQLite may return None)
    out: List[schemas.AdminCandidateStats] = []
    for r in rows:
        out.append(
            schemas.AdminCandidateStats(
                candidate_id=int(r.candidate_id),
                total_questions=int(r.total_questions or 0),
                total_feedback=int(r.total_feedback or 0),
                answered_questions=int(r.answered_questions or 0),
            )
        )
    return out


# ============================================================================
# ADMIN: MVP LEARNING PANEL
# ============================================================================


def _overview_counters(db: Session, *, candidate_id: int | None) -> schemas.MvpOverviewCounters:
    uq_users = db.query(func.count(func.distinct(models.BotUserRegistry.telegram_user_id)))
    active_users_q = db.query(func.count(func.distinct(models.BotUserRegistry.telegram_user_id)))
    active_cutoff = datetime.utcnow() - timedelta(days=7)
    if candidate_id is not None:
        uq_users = uq_users.filter(models.BotUserRegistry.candidate_id == int(candidate_id))
        active_users_q = active_users_q.filter(
            models.BotUserRegistry.candidate_id == int(candidate_id),
            models.BotUserRegistry.last_seen_at >= active_cutoff,
        )
    else:
        active_users_q = active_users_q.filter(models.BotUserRegistry.last_seen_at >= active_cutoff)

    total_users = int(uq_users.scalar() or 0)
    active_users = int(active_users_q.scalar() or 0)

    q = db.query(func.count(models.BotSubmission.id)).filter(models.BotSubmission.type == "QUESTION")
    q_ans = db.query(func.count(models.BotSubmission.id)).filter(
        models.BotSubmission.type == "QUESTION",
        func.upper(func.coalesce(models.BotSubmission.status, "")) == "ANSWERED",
    )
    fb = db.query(func.count(models.BotSubmission.id)).filter(models.BotSubmission.type == "FEEDBACK")
    leads = db.query(func.count(models.BotSubmission.id)).filter(models.BotSubmission.type == "BOT_REQUEST")
    if candidate_id is not None:
        q = q.filter(models.BotSubmission.candidate_id == int(candidate_id))
        q_ans = q_ans.filter(models.BotSubmission.candidate_id == int(candidate_id))
        fb = fb.filter(models.BotSubmission.candidate_id == int(candidate_id))
        leads = leads.filter(models.BotSubmission.candidate_id == int(candidate_id))

    total_questions = int(q.scalar() or 0)
    answered_questions = int(q_ans.scalar() or 0)
    total_comments = int(fb.scalar() or 0)
    total_leads = int(leads.scalar() or 0)

    commitments = db.query(func.count(models.BotCommitment.id))
    if candidate_id is not None:
        commitments = commitments.filter(models.BotCommitment.candidate_id == int(candidate_id))
    total_commitments = int(commitments.scalar() or 0)

    return schemas.MvpOverviewCounters(
        total_users=total_users,
        active_users=active_users,
        total_questions=total_questions,
        answered_questions=answered_questions,
        total_comments=total_comments,
        total_commitments=total_commitments,
        total_leads=total_leads,
    )


def _parse_iso_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        s = str(value).strip()
        if not s:
            return None
        if "T" not in s:
            return datetime.fromisoformat(s + "T00:00:00")
        return datetime.fromisoformat(s.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def _log_export_action(db: Session, *, admin_id: int, export_type: str, filters: dict | None = None) -> None:
    try:
        db.add(models.AdminExportLog(admin_id=int(admin_id), export_type=str(export_type), filters=filters or None))
        db.commit()
    except Exception:
        db.rollback()


@app.get("/api/admin/mvp/overview", response_model=schemas.MvpOverviewResponse)
def admin_mvp_overview(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    global_counters = _overview_counters(db, candidate_id=None)
    candidates = (
        db.query(models.User)
        .filter(models.User.role == "CANDIDATE")
        .order_by(models.User.id.asc())
        .all()
    )
    per_candidate: list[schemas.MvpRepresentativeOverview] = []
    for c in candidates:
        per_candidate.append(
            schemas.MvpRepresentativeOverview(
                candidate_id=int(c.id),
                name=getattr(c, "full_name", None) or getattr(c, "username", None),
                counters=_overview_counters(db, candidate_id=int(c.id)),
            )
        )
    return schemas.MvpOverviewResponse(global_counters=global_counters, per_candidate=per_candidate)


@app.get("/api/admin/mvp/behavior", response_model=schemas.BehaviorStatsResponse)
def admin_mvp_behavior_stats(
    candidate_id: Optional[int] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    q = db.query(models.BotBehaviorCounter)
    if candidate_id is None:
        q = q.filter(models.BotBehaviorCounter.candidate_id.is_(None))
    else:
        q = q.filter(models.BotBehaviorCounter.candidate_id == int(candidate_id))
    rows = q.order_by(models.BotBehaviorCounter.event.asc()).all()
    items = [schemas.BehaviorCounterItem(event=r.event, count=int(r.count or 0)) for r in rows]
    return schemas.BehaviorStatsResponse(candidate_id=candidate_id, items=items)


@app.get("/api/admin/mvp/paths", response_model=schemas.FlowPathsResponse)
def admin_mvp_flow_paths(
    candidate_id: Optional[int] = None,
    limit: int = 20,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    limit = max(1, min(int(limit), 200))
    q = db.query(models.BotFlowPathCounter)
    if candidate_id is None:
        q = q.filter(models.BotFlowPathCounter.candidate_id.is_(None))
    else:
        q = q.filter(models.BotFlowPathCounter.candidate_id == int(candidate_id))
    rows = q.order_by(models.BotFlowPathCounter.count.desc()).limit(limit).all()
    items = [schemas.FlowPathItem(path=r.path, count=int(r.count or 0)) for r in rows]
    return schemas.FlowPathsResponse(candidate_id=candidate_id, items=items)


@app.get("/api/admin/mvp/questions", response_model=List[schemas.QuestionLearningItem])
def admin_mvp_questions(
    candidate_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    q = db.query(models.BotSubmission).filter(models.BotSubmission.type == "QUESTION")
    if candidate_id is not None:
        q = q.filter(models.BotSubmission.candidate_id == int(candidate_id))
    if status:
        q = q.filter(func.upper(func.coalesce(models.BotSubmission.status, "")) == str(status).strip().upper())
    return q.order_by(models.BotSubmission.id.desc()).all()


@app.get("/api/admin/mvp/commitments", response_model=List[schemas.CommitmentLearningItem])
def admin_mvp_commitments(
    candidate_id: Optional[int] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    q = db.query(models.BotCommitment)
    if candidate_id is not None:
        q = q.filter(models.BotCommitment.candidate_id == int(candidate_id))
    return q.order_by(models.BotCommitment.id.desc()).all()


@app.get("/api/admin/mvp/leads", response_model=List[schemas.LeadItem])
def admin_mvp_leads(
    candidate_id: Optional[int] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    q = db.query(models.BotSubmission).filter(models.BotSubmission.type == "BOT_REQUEST")
    if candidate_id is not None:
        q = q.filter(models.BotSubmission.candidate_id == int(candidate_id))
    return q.order_by(models.BotSubmission.id.desc()).all()


@app.get("/api/admin/mvp/ux-logs", response_model=List[schemas.UxLogItem])
def admin_mvp_ux_logs(
    candidate_id: Optional[int] = None,
    action: Optional[str] = None,
    limit: int = 200,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    limit = max(1, min(int(limit), 2000))
    q = db.query(models.BotUxLog)
    if candidate_id is not None:
        q = q.filter(models.BotUxLog.candidate_id == int(candidate_id))
    if action:
        q = q.filter(models.BotUxLog.action == str(action).strip())
    return q.order_by(models.BotUxLog.id.desc()).limit(limit).all()


@app.get("/api/admin/mvp/global-users", response_model=List[schemas.GlobalBotUserItem])
def admin_mvp_global_users(
    representative_id: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    interaction_type: Optional[str] = None,
    limit: int = 1000,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    limit = max(1, min(int(limit), 5000))
    q = db.query(models.BotUserRegistry)

    if representative_id is not None:
        q = q.filter(models.BotUserRegistry.candidate_id == int(representative_id))

    sd = _parse_iso_dt(start_date)
    ed = _parse_iso_dt(end_date)
    if sd is not None:
        q = q.filter(models.BotUserRegistry.first_seen_at >= sd)
    if ed is not None:
        q = q.filter(models.BotUserRegistry.last_seen_at <= ed)

    it = (interaction_type or "").strip().lower()
    if it == "question":
        q = q.filter(models.BotUserRegistry.asked_question == True)  # noqa: E712
    elif it == "comment":
        q = q.filter(models.BotUserRegistry.left_comment == True)  # noqa: E712
    elif it == "lead":
        q = q.filter(models.BotUserRegistry.became_lead == True)  # noqa: E712

    return q.order_by(models.BotUserRegistry.last_seen_at.desc()).limit(limit).all()


@app.get("/api/admin/mvp/global-users/export.xlsx")
def admin_mvp_global_users_export_xlsx(
    representative_id: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    interaction_type: Optional[str] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    rows = admin_mvp_global_users(
        representative_id=representative_id,
        start_date=start_date,
        end_date=end_date,
        interaction_type=interaction_type,
        limit=5000,
        db=db,
        current_user=current_user,
    )

    # Log export action (non-fatal)
    _log_export_action(
        db,
        admin_id=int(current_user.id),
        export_type="global_users",
        filters={
            "representative_id": representative_id,
            "start_date": start_date,
            "end_date": end_date,
            "interaction_type": interaction_type,
        },
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "users"

    headers = [
        "id",
        "user_id",
        "username",
        "first_name",
        "last_name",
        "phone",
        "platform",
        "representative_id",
        "bot_id",
        "candidate_name",
        "candidate_city",
        "candidate_province",
        "candidate_constituency",
        "chat_type",
        "first_interaction_at",
        "last_interaction_at",
        "total_interactions",
        "asked_question",
        "left_comment",
        "viewed_commitment",
        "became_lead",
        "selected_role",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                int(getattr(r, "id", 0) or 0),
                str(getattr(r, "telegram_user_id", "") or ""),
                str(getattr(r, "telegram_username", "") or ""),
                str(getattr(r, "first_name", "") or ""),
                str(getattr(r, "last_name", "") or ""),
                str(getattr(r, "phone", "") or ""),
                str(getattr(r, "platform", "TELEGRAM") or "TELEGRAM"),
                int(getattr(r, "candidate_id", 0) or 0),
                str(getattr(r, "candidate_bot_name", "") or ""),
                str(getattr(r, "candidate_name", "") or ""),
                str(getattr(r, "candidate_city", "") or ""),
                str(getattr(r, "candidate_province", "") or ""),
                str(getattr(r, "candidate_constituency", "") or ""),
                str(getattr(r, "chat_type", "") or ""),
                getattr(r, "first_seen_at", None),
                getattr(r, "last_seen_at", None),
                int(getattr(r, "total_interactions", 0) or 0),
                bool(getattr(r, "asked_question", False)),
                bool(getattr(r, "left_comment", False)),
                bool(getattr(r, "viewed_commitment", False)),
                bool(getattr(r, "became_lead", False)),
                str(getattr(r, "selected_role", "") or ""),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"global_users_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# =====================
# Monitoring (Super Admin only)
# =====================


@app.get("/api/admin/monitoring/errors", response_model=List[schemas.TechnicalErrorItem])
def admin_monitoring_errors(
    representative_id: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 500,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    limit = max(1, min(int(limit), 5000))
    q = db.query(models.TechnicalErrorLog)
    if representative_id is not None:
        q = q.filter(models.TechnicalErrorLog.candidate_id == int(representative_id))
    sd = _parse_iso_dt(start_date)
    ed = _parse_iso_dt(end_date)
    if sd is not None:
        q = q.filter(models.TechnicalErrorLog.created_at >= sd)
    if ed is not None:
        q = q.filter(models.TechnicalErrorLog.created_at <= ed)
    return q.order_by(models.TechnicalErrorLog.id.desc()).limit(limit).all()


@app.get("/api/admin/monitoring/errors/export.xlsx")
def admin_monitoring_errors_export_xlsx(
    representative_id: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    rows = admin_monitoring_errors(
        representative_id=representative_id,
        start_date=start_date,
        end_date=end_date,
        limit=5000,
        db=db,
        current_user=current_user,
    )
    _log_export_action(
        db,
        admin_id=int(current_user.id),
        export_type="technical_errors",
        filters={"representative_id": representative_id, "start_date": start_date, "end_date": end_date},
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "errors"
    ws.append(
        [
            "error_id",
            "timestamp",
            "service_name",
            "error_type",
            "error_message",
            "user_id",
            "representative_id",
            "state",
        ]
    )
    for r in rows:
        ws.append(
            [
                int(getattr(r, "id", 0) or 0),
                getattr(r, "created_at", None),
                str(getattr(r, "service_name", "") or ""),
                str(getattr(r, "error_type", "") or ""),
                str(getattr(r, "error_message", "") or ""),
                str(getattr(r, "telegram_user_id", "") or ""),
                getattr(r, "candidate_id", None),
                str(getattr(r, "state", "") or ""),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"monitoring_errors_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/admin/monitoring/ux-logs", response_model=List[schemas.MonitoringUxLogItem])
def admin_monitoring_ux_logs(
    representative_id: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    action: Optional[str] = None,
    limit: int = 1000,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    limit = max(1, min(int(limit), 5000))
    q = db.query(models.BotUxLog)
    if representative_id is not None:
        q = q.filter(models.BotUxLog.candidate_id == int(representative_id))
    sd = _parse_iso_dt(start_date)
    ed = _parse_iso_dt(end_date)
    if sd is not None:
        q = q.filter(models.BotUxLog.created_at >= sd)
    if ed is not None:
        q = q.filter(models.BotUxLog.created_at <= ed)
    if action:
        q = q.filter(models.BotUxLog.action == str(action).strip())
    return q.order_by(models.BotUxLog.id.desc()).limit(limit).all()


@app.get("/api/admin/monitoring/ux-logs/export.xlsx")
def admin_monitoring_ux_logs_export_xlsx(
    representative_id: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    action: Optional[str] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    rows = admin_monitoring_ux_logs(
        representative_id=representative_id,
        start_date=start_date,
        end_date=end_date,
        action=action,
        limit=5000,
        db=db,
        current_user=current_user,
    )
    _log_export_action(
        db,
        admin_id=int(current_user.id),
        export_type="ux_logs",
        filters={
            "representative_id": representative_id,
            "start_date": start_date,
            "end_date": end_date,
            "action": action,
        },
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "ux_logs"
    ws.append(
        [
            "log_id",
            "timestamp",
            "user_id",
            "representative_id",
            "current_state",
            "action",
            "expected_action",
        ]
    )
    for r in rows:
        ws.append(
            [
                int(getattr(r, "id", 0) or 0),
                getattr(r, "created_at", None),
                str(getattr(r, "telegram_user_id", "") or ""),
                int(getattr(r, "candidate_id", 0) or 0),
                str(getattr(r, "state", "") or ""),
                str(getattr(r, "action", "") or ""),
                str(getattr(r, "expected_action", "") or ""),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"monitoring_ux_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/admin/monitoring/flow-drops", response_model=List[schemas.FlowDropItem])
def admin_monitoring_flow_drops(
    representative_id: Optional[int] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    q = db.query(models.BotFlowDropCounter)
    if representative_id is not None:
        q = q.filter(models.BotFlowDropCounter.candidate_id == int(representative_id))
    return q.order_by(models.BotFlowDropCounter.updated_at.desc()).all()


@app.get("/api/admin/monitoring/flow-drops/export.xlsx")
def admin_monitoring_flow_drops_export_xlsx(
    representative_id: Optional[int] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    rows = admin_monitoring_flow_drops(representative_id=representative_id, db=db, current_user=current_user)
    _log_export_action(
        db,
        admin_id=int(current_user.id),
        export_type="flow_drop_stats",
        filters={"representative_id": representative_id},
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "flow_drops"
    ws.append(["id", "representative_id", "flow_type", "started_count", "completed_count", "abandoned_count", "updated_at"])
    for r in rows:
        ws.append(
            [
                int(getattr(r, "id", 0) or 0),
                getattr(r, "candidate_id", None),
                str(getattr(r, "flow_type", "") or ""),
                int(getattr(r, "started_count", 0) or 0),
                int(getattr(r, "completed_count", 0) or 0),
                int(getattr(r, "abandoned_count", 0) or 0),
                getattr(r, "updated_at", None),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"monitoring_flow_drops_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/admin/monitoring/health-checks", response_model=List[schemas.HealthCheckItem])
def admin_monitoring_health_checks(
    representative_id: Optional[int] = None,
    check_type: Optional[str] = None,
    limit: int = 500,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_super_admin_user),
):
    limit = max(1, min(int(limit), 5000))
    q = db.query(models.BotHealthCheck)
    if representative_id is not None:
        q = q.filter(models.BotHealthCheck.candidate_id == int(representative_id))
    if check_type:
        q = q.filter(models.BotHealthCheck.check_type == str(check_type).strip())
    return q.order_by(models.BotHealthCheck.id.desc()).limit(limit).all()


@app.put("/api/candidates/me/feedback/{submission_id}", response_model=schemas.FeedbackSubmission)
def update_my_feedback_submission(
    submission_id: int,
    payload: schemas.FeedbackSubmissionUpdate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.role != "CANDIDATE":
        raise HTTPException(status_code=403, detail="Access denied")

    submission = (
        db.query(models.BotSubmission)
        .filter(
            models.BotSubmission.id == submission_id,
            models.BotSubmission.candidate_id == current_user.id,
            models.BotSubmission.type == "FEEDBACK",
        )
        .first()
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")

    data = payload.model_dump(exclude_unset=True)

    if "tag" in data:
        raw = data.get("tag")
        if raw is None:
            submission.tag = None
        else:
            trimmed = str(raw).strip()
            submission.tag = trimmed or None

    if "status" in data:
        raw_status = (data.get("status") or "").strip().upper()
        if raw_status not in {"NEW", "REVIEWED"}:
            raise HTTPException(status_code=422, detail={"message": "Invalid status"})
        submission.status = raw_status

    db.add(submission)
    db.commit()
    db.refresh(submission)
    return submission


@app.get("/api/candidates/me/feedback/stats", response_model=schemas.FeedbackStatsResponse)
def get_my_feedback_stats(
    days: int = 7,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.role != "CANDIDATE":
        raise HTTPException(status_code=403, detail="Access denied")

    if days not in {7, 30}:
        raise HTTPException(status_code=422, detail={"message": "days must be 7 or 30"})

    since = datetime.utcnow() - timedelta(days=days)
    rows = (
        db.query(models.BotSubmission)
        .filter(
            models.BotSubmission.candidate_id == current_user.id,
            models.BotSubmission.type == "FEEDBACK",
            models.BotSubmission.created_at >= since,
        )
        .all()
    )

    total = len(rows)
    counts: dict[str, int] = {}
    for r in rows:
        tag = (r.tag or "").strip() or "سایر"
        counts[tag] = counts.get(tag, 0) + 1

    items = []
    for tag, count in sorted(counts.items(), key=lambda kv: kv[1], reverse=True):
        percent = round((count / total * 100.0), 2) if total else 0.0
        items.append({"tag": tag, "count": int(count), "percent": float(percent)})

    return {"days": days, "total": total, "items": items}


# ============================================================================
# QUESTIONS (PUBLIC Q&A) ENDPOINTS (MVP)
# ============================================================================

@app.get("/api/candidates/me/questions", response_model=List[schemas.QuestionSubmission])
def get_my_question_submissions(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.role != "CANDIDATE":
        raise HTTPException(status_code=403, detail="Access denied")

    return (
        db.query(models.BotSubmission)
        .filter(
            models.BotSubmission.candidate_id == current_user.id,
            models.BotSubmission.type == "QUESTION",
        )
        .order_by(models.BotSubmission.id.desc())
        .all()
    )


@app.put("/api/candidates/me/questions/{submission_id}/answer", response_model=schemas.QuestionSubmission)
def answer_my_question_submission(
    submission_id: int,
    payload: schemas.QuestionSubmissionAnswer,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.role != "CANDIDATE":
        raise HTTPException(status_code=403, detail="Access denied")

    submission = (
        db.query(models.BotSubmission)
        .filter(
            models.BotSubmission.id == submission_id,
            models.BotSubmission.candidate_id == current_user.id,
            models.BotSubmission.type == "QUESTION",
        )
        .first()
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")

    if (submission.status or "").upper() == "ANSWERED":
        raise HTTPException(status_code=409, detail={"message": "Answer is immutable"})

    answer_text = (payload.answer_text or "").strip()
    if not answer_text:
        raise HTTPException(status_code=422, detail={"message": "answer_text is required"})
    if len(answer_text) > 2000:
        raise HTTPException(status_code=422, detail={"message": "answer_text is too long"})

    submission.answer = answer_text
    submission.status = "ANSWERED"
    submission.is_public = True
    submission.answered_at = datetime.utcnow()

    # Optional metadata
    if getattr(payload, "topic", None) is not None:
        submission.topic = (payload.topic or "").strip() or None
    if getattr(payload, "is_featured", None) is not None:
        submission.is_featured = bool(payload.is_featured)

    db.add(submission)
    db.commit()
    db.refresh(submission)

    # Best-effort announcement to candidate group/channel (notification-only).
    _notify_question_answer_published(candidate=current_user, submission=submission)
    return submission


@app.put("/api/candidates/me/questions/{submission_id}/meta", response_model=schemas.QuestionSubmission)
def update_my_question_submission_meta(
    submission_id: int,
    payload: schemas.QuestionSubmissionMeta,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.role != "CANDIDATE":
        raise HTTPException(status_code=403, detail="Access denied")

    submission = (
        db.query(models.BotSubmission)
        .filter(
            models.BotSubmission.id == submission_id,
            models.BotSubmission.candidate_id == current_user.id,
            models.BotSubmission.type == "QUESTION",
        )
        .first()
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")

    if payload.topic is not None:
        submission.topic = (payload.topic or "").strip() or None
    if payload.is_featured is not None:
        submission.is_featured = bool(payload.is_featured)

    db.add(submission)
    db.commit()
    db.refresh(submission)
    return submission


@app.put("/api/candidates/me/questions/{submission_id}/reject", response_model=schemas.QuestionSubmission)
def reject_my_question_submission(
    submission_id: int,
    _payload: schemas.QuestionSubmissionReject | None = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if current_user.role != "CANDIDATE":
        raise HTTPException(status_code=403, detail="Access denied")

    submission = (
        db.query(models.BotSubmission)
        .filter(
            models.BotSubmission.id == submission_id,
            models.BotSubmission.candidate_id == current_user.id,
            models.BotSubmission.type == "QUESTION",
        )
        .first()
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")

    if (submission.status or "").upper() == "ANSWERED":
        raise HTTPException(status_code=409, detail={"message": "Answer is immutable"})

    submission.status = "REJECTED"
    submission.is_public = False

    db.add(submission)
    db.commit()
    db.refresh(submission)
    return submission


# ============================================================================
# COMMITMENTS (PUBLIC DIGITAL CONTRACTS) ENDPOINTS (STRICT)
# ============================================================================


@app.get("/api/candidates/me/commitments/terms/acceptance", response_model=Optional[schemas.CommitmentTermsAcceptanceOut])
def get_commitment_terms_acceptance(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    _require_candidate(current_user)
    row = (
        db.query(models.CommitmentTermsAcceptance)
        .filter(models.CommitmentTermsAcceptance.representative_id == int(current_user.id))
        .first()
    )
    return row


@app.post("/api/candidates/me/commitments/terms/accept", response_model=schemas.CommitmentTermsAcceptanceOut)
def accept_commitment_terms(
    request: Request,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    _require_candidate(current_user)

    existing = (
        db.query(models.CommitmentTermsAcceptance)
        .filter(models.CommitmentTermsAcceptance.representative_id == int(current_user.id))
        .first()
    )
    if existing is not None:
        return existing

    ip_address = None
    try:
        ip_address = getattr(getattr(request, "client", None), "host", None)
    except Exception:
        ip_address = None
    user_agent = (request.headers.get("user-agent") or "").strip() or None

    row = models.CommitmentTermsAcceptance(
        representative_id=int(current_user.id),
        accepted_at=datetime.utcnow(),
        ip_address=ip_address,
        user_agent=user_agent,
        version=COMMITMENT_TERMS_VERSION,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.get("/api/candidates/me/commitments", response_model=List[schemas.CommitmentOut])
def list_my_commitments(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    _require_candidate(current_user)
    rows = (
        db.query(models.BotCommitment)
        .options(joinedload(models.BotCommitment.progress_logs))
        .filter(models.BotCommitment.candidate_id == int(current_user.id))
        .order_by(models.BotCommitment.id.desc())
        .all()
    )
    # Legacy SQLite DBs may contain NULL status_updated_at (column added later).
    # Our response schema requires a datetime, so fill it defensively.
    touched = False
    now = datetime.utcnow()
    for r in rows:
        if getattr(r, "status_updated_at", None) is None:
            r.status_updated_at = getattr(r, "published_at", None) or getattr(r, "created_at", None) or now
            touched = True
    if touched:
        db.commit()
    return rows


@app.post("/api/candidates/me/commitments", response_model=schemas.CommitmentOut)
def create_commitment_draft(
    payload: schemas.CommitmentCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    _require_candidate(current_user)

    accepted = (
        db.query(models.CommitmentTermsAcceptance)
        .filter(models.CommitmentTermsAcceptance.representative_id == int(current_user.id))
        .first()
    )
    if accepted is None:
        raise HTTPException(status_code=403, detail={"message": "Commitment terms not accepted"})

    title = (payload.title or "").strip()
    description = (payload.description or "").strip()
    category = _validate_commitment_category(payload.category)

    if not title:
        raise HTTPException(status_code=422, detail={"message": "title is required"})
    if len(title) > 120:
        raise HTTPException(status_code=422, detail={"message": "title is too long"})
    if not description:
        raise HTTPException(status_code=422, detail={"message": "description is required"})
    if len(description) > 5000:
        raise HTTPException(status_code=422, detail={"message": "description is too long"})

    row = models.BotCommitment(
        candidate_id=int(current_user.id),
        title=title,
        body=description,
        category=category,
        status="draft",
        status_updated_at=datetime.utcnow(),
        locked=False,
        published_at=None,
        created_at=datetime.utcnow(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.put("/api/candidates/me/commitments/{commitment_id}", response_model=schemas.CommitmentOut)
def update_commitment_draft(
    commitment_id: int,
    payload: schemas.CommitmentUpdateDraft,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    _require_candidate(current_user)

    row = (
        db.query(models.BotCommitment)
        .options(joinedload(models.BotCommitment.progress_logs))
        .filter(models.BotCommitment.id == int(commitment_id), models.BotCommitment.candidate_id == int(current_user.id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Commitment not found")

    is_published = bool(getattr(row, "published_at", None)) or bool(getattr(row, "locked", False))
    if is_published or _normalize_commitment_status(getattr(row, "status", None)) != "draft":
        _log_commitment_security_event(
            db=db,
            representative_id=int(current_user.id),
            error_type="CommitmentImmutableViolation",
            message=f"Attempted to edit commitment fields after publish. commitment_id={commitment_id}",
        )
        raise HTTPException(status_code=403, detail={"message": "Commitment is immutable after publish"})

    data = payload.model_dump(exclude_unset=True)
    if "title" in data:
        v = (data.get("title") or "").strip()
        if not v:
            raise HTTPException(status_code=422, detail={"message": "title is required"})
        if len(v) > 120:
            raise HTTPException(status_code=422, detail={"message": "title is too long"})
        row.title = v

    if "description" in data:
        v = (data.get("description") or "").strip()
        if not v:
            raise HTTPException(status_code=422, detail={"message": "description is required"})
        if len(v) > 5000:
            raise HTTPException(status_code=422, detail={"message": "description is too long"})
        row.body = v

    if "category" in data:
        row.category = _validate_commitment_category(data.get("category"))

    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.post("/api/candidates/me/commitments/{commitment_id}/publish", response_model=schemas.CommitmentOut)
def publish_commitment(
    commitment_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    _require_candidate(current_user)

    row = (
        db.query(models.BotCommitment)
        .options(joinedload(models.BotCommitment.progress_logs))
        .filter(models.BotCommitment.id == int(commitment_id), models.BotCommitment.candidate_id == int(current_user.id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Commitment not found")

    if getattr(row, "published_at", None) is not None or bool(getattr(row, "locked", False)):
        raise HTTPException(status_code=409, detail={"message": "Commitment already published"})

    if _normalize_commitment_status(getattr(row, "status", None)) != "draft":
        raise HTTPException(status_code=409, detail={"message": "Only draft commitments can be published"})

    # Enforce required fields before publish
    if not (getattr(row, "title", None) or "").strip() or not (getattr(row, "body", None) or "").strip():
        raise HTTPException(status_code=422, detail={"message": "title and description are required"})
    if not (getattr(row, "category", None) or "").strip():
        raise HTTPException(status_code=422, detail={"message": "category is required"})

    now = datetime.utcnow()
    row.published_at = now
    row.locked = True
    row.status = "active"
    row.status_updated_at = now

    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.post("/api/candidates/me/commitments/{commitment_id}/status", response_model=schemas.CommitmentOut)
def update_commitment_status(
    commitment_id: int,
    payload: schemas.CommitmentUpdateStatus,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    _require_candidate(current_user)

    row = (
        db.query(models.BotCommitment)
        .options(joinedload(models.BotCommitment.progress_logs))
        .filter(models.BotCommitment.id == int(commitment_id), models.BotCommitment.candidate_id == int(current_user.id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Commitment not found")

    if getattr(row, "published_at", None) is None or not bool(getattr(row, "locked", False)):
        raise HTTPException(status_code=403, detail={"message": "Only published commitments can change status"})

    next_status = _validate_commitment_status_after_publish(payload.status)
    now = datetime.utcnow()
    row.status = next_status
    row.status_updated_at = now

    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.post("/api/candidates/me/commitments/{commitment_id}/progress", response_model=schemas.CommitmentOut)
def add_commitment_progress_log(
    commitment_id: int,
    payload: schemas.CommitmentAddProgress,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    _require_candidate(current_user)

    row = (
        db.query(models.BotCommitment)
        .options(joinedload(models.BotCommitment.progress_logs))
        .filter(models.BotCommitment.id == int(commitment_id), models.BotCommitment.candidate_id == int(current_user.id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Commitment not found")

    if getattr(row, "published_at", None) is None or not bool(getattr(row, "locked", False)):
        raise HTTPException(status_code=403, detail={"message": "Progress logs are only allowed after publish"})

    note = (payload.note or "").strip()
    if not note:
        raise HTTPException(status_code=422, detail={"message": "note is required"})
    if len(note) > 1000:
        raise HTTPException(status_code=422, detail={"message": "note is too long"})

    log_row = models.CommitmentProgressLog(commitment_id=int(row.id), note=note, created_at=datetime.utcnow())
    db.add(log_row)
    db.commit()

    # refresh parent with logs
    row = (
        db.query(models.BotCommitment)
        .options(joinedload(models.BotCommitment.progress_logs))
        .filter(models.BotCommitment.id == int(commitment_id), models.BotCommitment.candidate_id == int(current_user.id))
        .first()
    )
    return row


@app.delete("/api/candidates/me/commitments/{commitment_id}")
def delete_commitment_draft(
    commitment_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    _require_candidate(current_user)

    row = (
        db.query(models.BotCommitment)
        .filter(models.BotCommitment.id == int(commitment_id), models.BotCommitment.candidate_id == int(current_user.id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Commitment not found")

    if getattr(row, "published_at", None) is not None or bool(getattr(row, "locked", False)):
        _log_commitment_security_event(
            db=db,
            representative_id=int(current_user.id),
            error_type="CommitmentDeleteForbidden",
            message=f"Attempted to delete commitment after publish. commitment_id={commitment_id}",
        )
        raise HTTPException(status_code=403, detail={"message": "Published commitments cannot be deleted"})

    if _normalize_commitment_status(getattr(row, "status", None)) != "draft":
        raise HTTPException(status_code=403, detail={"message": "Only draft commitments can be deleted"})

    db.delete(row)
    db.commit()
    return {"message": "deleted"}

# ============================================================================
# ANNOUNCEMENTS ENDPOINTS
# ============================================================================

@app.get("/api/announcements", response_model=List[schemas.Announcement])
def get_announcements(db: Session = Depends(database.get_db)):
    return db.query(models.Announcement).order_by(models.Announcement.id.desc()).all()

@app.post("/api/announcements", response_model=schemas.Announcement)
def create_announcement(
    announcement: schemas.AnnouncementCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_admin_user),
):
    now_jalali = jdatetime.datetime.now()
    created_at_jalali = now_jalali.strftime("%Y/%m/%d %H:%M:%S")

    new_announcement = models.Announcement(
        title=announcement.title,
        content=announcement.content,
        attachments=announcement.attachments,
        created_at_jalali=created_at_jalali
    )
    db.add(new_announcement)
    db.commit()
    db.refresh(new_announcement)
    return new_announcement

@app.get("/")
async def root():
    return {"message": "Election Manager API"}

@app.get("/health")
async def health():
    return {"status": "healthy"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
